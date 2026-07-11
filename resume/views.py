import os
import email
import traceback
from urllib import request
from django.conf import settings

from django.contrib.auth.hashers import make_password, check_password
from django.shortcuts import render, redirect,get_object_or_404
from django.db.models import Q
from django.http import HttpResponse
from django.http import JsonResponse
from django.contrib import messages
from django.db.models import Avg, Max, Min, Count
from django.db.models.functions import ExtractMonth

import resume
from resume import admin
from .models import Register, Resume, Contact
from django.core.paginator import Paginator
from PyPDF2 import PdfReader
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from openpyxl import Workbook
from django.utils import timezone
from openpyxl.styles import Font, Alignment,PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import csv

# ===========================
# Home
# ===========================

def home(request):
    username = request.session.get("user_name")

    return render(request, "home.html", {
        "username": username
    })


# ===========================
# Register
# ===========================

def register(request):

    if request.method == "POST":
        
        print("REGISTER POST")
        print(request.POST)
        print(request.FILES)

        fullname = request.POST.get("fullName", "").strip()
        email = request.POST.get("email", "").strip().lower()
        password = request.POST.get("password", "")
        confirm_password = request.POST.get("confirm_password", "")
        profile_photo = request.FILES.get("profile_photo")

        # Full Name Validation
        if not fullname:
            messages.error(request, "Full Name is required.")
            return render(request, "register.html")

        # Email Validation
        if not email:
            messages.error(request, "Email is required.")
            return render(request, "register.html")

        # Password Validation
        if not password:
            messages.error(request, "Password is required.")
            return render(request, "register.html")

        if len(password) < 8:
            messages.error(request, "Password must be at least 8 characters.")
            return render(request, "register.html")

        # Confirm Password
        if password != confirm_password:
            messages.error(request, "Password and Confirm Password do not match.")
            return render(request, "register.html")

        # Email Already Exists
        if Register.objects.filter(email=email).exists():
            messages.error(request, "Email already registered.")
            return render(request, "register.html")

        # Profile Photo Validation
        if profile_photo:

            allowed_extensions = ["jpg", "jpeg", "png"]

            extension = profile_photo.name.split(".")[-1].lower()

            if extension not in allowed_extensions:
                messages.error(
                    request,
                    "Only JPG, JPEG and PNG images are allowed."
                )
                return render(request, "register.html")

        # Create User
        print("BEFORE CREATE")
        Register.objects.create(
            fullname=fullname,
            email=email,
            password=make_password(password),
            profile_photo=profile_photo
        )
        print("USER CREATED")

        messages.success(request, "Registration Successful.")

        return redirect("login")

    return render(request, "register.html")

# ===========================
# Login
# ===========================

def login(request):

    # Already logged in as normal user
    if request.session.get("user_id") and not request.session.get("is_admin"):
        return redirect("dashboard")

    # Already logged in as admin
    if request.session.get("is_admin"):
        return redirect("admin_dashboard")

    if request.method == "POST":
        
        print("LOGIN POST")
        print(request.POST)

        email = request.POST.get("email", "").strip().lower()
        password = request.POST.get("password", "")

        if not email or not password:
            messages.error(request, "Email and Password are required.")
            return render(request, "login.html")

        user = Register.objects.filter(email=email).first()
        
        print("USER =", user)
        if not user:
            print("USER NOT FOUND")
            print("EMAIL =", email)


        
            messages.error(request, "Invalid Email or Password.")
            return render(request, "login.html")
        print("DB PASSWORD =", user.password)
        print("CHECK =", check_password(password, user.password))


        # Password Verification
        if not check_password(password, user.password):
            messages.error(request, "Invalid Email or Password.")
            return render(request, "login.html")

        # Admin cannot login from User Login page
        if user.is_admin:
            messages.error(
                request,
                "Please login from Admin Login."
            )
            return redirect("admin_login")

        # User Session
        request.session["user_id"] = user.id
        request.session["user_name"] = user.fullname
        request.session["is_admin"] = False

        messages.success(request, "Login Successful.")

        return redirect("dashboard")

    return render(request, "login.html")

# ===========================
# Logout
# ===========================

def logout(request):
    request.session.flush()
    return redirect("login")


# ===========================
# Upload Resume
# ===========================

def upload_resume(request):

    if "user_id" not in request.session:
        return redirect("login")

    if request.method == "POST":

        resume_file = request.FILES.get("resume")

        # File Required
        if not resume_file:
            messages.error(
                request,
                "Please select a PDF resume."
            )
            return render(request, "upload_resume.html")

        # File Size Validation
        if resume_file.size > 5 * 1024 * 1024:
            messages.error(
                request,
                "Maximum file size is 5 MB."
            )
            return render(request, "upload_resume.html")

        # PDF Validation
        if not resume_file.name.lower().endswith(".pdf"):
            messages.error(
                request,
                "Only PDF files are allowed."
            )
            return render(request, "upload_resume.html")

        # -------------------------
        # Read PDF BEFORE Uploading
        # -------------------------

        try:

            resume_file.seek(0)

            reader = PdfReader(resume_file)

            text = ""

            for page in reader.pages:

                extracted = page.extract_text()

                if extracted:
                    text += extracted

            resume_file.seek(0)

        except Exception as e:

            print("PDF Error:", repr(e))

            messages.error(
                request,
                "Invalid or Corrupted PDF."
            )

            return render(
                request,
                "upload_resume.html"
            )

        # -------------------------
        # User
        # -------------------------

        user = Register.objects.get(
            id=request.session["user_id"]
        )

        # -------------------------
        # ATS Analysis
        # -------------------------

        lower_text = text.lower()

        keywords = [
            "python",
            "django",
            "sql",
            "html",
            "css",
            "javascript",
            "machine learning",
            "ai",
            "react",
            "git",
            "github",
            "docker",
            "aws",
            "flask",
        ]

        skills_found = []
        missing_skills = []

        score = 0

        for skill in keywords:

            if skill in lower_text:
                skills_found.append(skill.title())
                score += 10
            else:
                missing_skills.append(skill.title())

        score = min(score, 100)

        suggestions = []

        if "react" not in lower_text:
            suggestions.append("Learn React")

        if "github" not in lower_text:
            suggestions.append("Add GitHub Profile")

        if "docker" not in lower_text:
            suggestions.append("Learn Docker")

        if "aws" not in lower_text:
            suggestions.append("Learn AWS Cloud")

        if "internship" not in lower_text:
            suggestions.append("Add Internship Experience")

        if "project" not in lower_text:
            suggestions.append("Add Personal Projects")

        if "certification" not in lower_text:
            suggestions.append("Add Certifications")

        # -------------------------
        # Save Resume
        # -------------------------

        

        try:

            print("Before Resume Create")

            resume = Resume.objects.create(
                user=user,
                resume=resume_file,
                extracted_text=text,
                ats_score=score,
                skills_found=", ".join(skills_found),
                missing_skills=", ".join(missing_skills),
                ai_suggestion=", ".join(suggestions),
            )
            print("URL =", resume.resume.url)
            print("NAME =", resume.resume.name)
            print("CLASS =", resume.resume.storage.__class__)
            print("STORAGE =", settings.DEFAULT_FILE_STORAGE)
            print("Resume Saved Successfully")

        except Exception as e:

            print("===== DATABASE ERROR =====")
            print(str(e))
            print(traceback.format_exc())

            raise

        messages.success(
            request,
            "Resume Uploaded Successfully."
        )

        print("Redirecting Result Page")

        return redirect(
            "result",
            id=resume.id
        )

    return render(
        request,
        "upload_resume.html"
    )
# ===========================
# Dashboard
# ===========================

def dashboard(request):

    if "user_id" not in request.session:
        return redirect("login")

    user = Register.objects.get(id=request.session["user_id"])

    resumes = Resume.objects.filter(user=user).order_by("-uploaded_at")

    latest_five = resumes[:5]

    total_resume = resumes.count()

    latest_resume = resumes.first()

    skills_found_list = []
    missing_skills_list = []

    if latest_resume:

        if latest_resume.skills_found:
            skills_found_list = [
                skill.strip()
                for skill in latest_resume.skills_found.split(",")
                if skill.strip()
            ]

        if latest_resume.missing_skills:
            missing_skills_list = [
                skill.strip()
                for skill in latest_resume.missing_skills.split(",")
                if skill.strip()
            ]

    highest_score = 0
    average_score = 0
    ai_rating = "No Resume"

    if resumes.exists():

        highest_score = max(r.ats_score for r in resumes)

        average_score = round(
            sum(r.ats_score for r in resumes) / resumes.count(),
            2
        )

        if highest_score >= 80:
            ai_rating = "Excellent ⭐⭐⭐⭐⭐"
        elif highest_score >= 60:
            ai_rating = "Good ⭐⭐⭐⭐"
        elif highest_score >= 40:
            ai_rating = "Average ⭐⭐⭐"
        else:
            ai_rating = "Needs Improvement ⭐⭐"

    context = {
        "user": user,
        "resumes": resumes,
        "latest_five": latest_five,
        "total_resume": total_resume,
        "latest_resume": latest_resume,
        "highest_score": highest_score,
        "average_score": average_score,
        "ai_rating": ai_rating,
        "skills_found_list": skills_found_list,
        "missing_skills_list": missing_skills_list,
    }

    return render(request, "dashboard.html", context)
# ===========================
# Result Page
# ===========================

def result(request, id):

    if "user_id" not in request.session:
        return redirect("login")

    user = Register.objects.get(id=request.session["user_id"])

    try:
        resume = Resume.objects.get(id=id, user=user)
    except Resume.DoesNotExist:
        return redirect("history")

    return render(request, "result.html", {
        "resume": resume,
        "user": user,
    })


# ===========================
# Resume History
# ===========================

def history(request):

    if "user_id" not in request.session:
        return redirect("login")

    user_id = request.session["user_id"]

    search = request.GET.get("search", "").strip()

    resumes = Resume.objects.filter(user_id=user_id).order_by("-uploaded_at")

    if search:
        resumes = resumes.filter(
            Q(resume__icontains=search) |
            Q(extracted_text__icontains=search) |
            Q(skills_found__icontains=search) |
            Q(missing_skills__icontains=search) |
            Q(ai_suggestion__icontains=search) |
            Q(ats_score__icontains=search)
        )

    paginator = Paginator(resumes, 10)

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "search": search,
    }

    return render(request, "history.html", context)


# ==========================================
# Export CSV (Professional)
# ==========================================

def export_csv(request):

    if "user_id" not in request.session:
        return redirect("login")

    user_id = request.session["user_id"]

    resumes = (
        Resume.objects
        .filter(user_id=user_id)
        .order_by("-uploaded_at")
    )

    response = HttpResponse(
        content_type="text/csv; charset=utf-8"
    )

    response["Content-Disposition"] = (
        'attachment; filename="resume_history.csv"'
    )

    # Excel UTF-8 Support
    response.write("\ufeff")

    writer = csv.writer(
        response,
        delimiter=",",
        quotechar='"',
        quoting=csv.QUOTE_MINIMAL,
        lineterminator="\n"
    )

    writer.writerow([
        "Resume",
        "Upload Date",
        "ATS Score",
        "Skills Found",
        "Missing Skills",
        "AI Suggestion",
    ])

    for resume in resumes:

        resume_name = ""

        if resume.resume:
            resume_name = resume.resume.name.split("/")[-1]

        writer.writerow([

            resume_name,

            resume.uploaded_at.strftime("%d-%m-%Y %I:%M %p")
            if resume.uploaded_at else "-",

            f"{resume.ats_score}%",

            resume.skills_found or "-",

            resume.missing_skills or "-",

            resume.ai_suggestion or "-",

        ])

    return response
# ==========================================
# Export Excel (Professional)
# ==========================================

def export_excel(request):

    if "user_id" not in request.session:
        return redirect("login")

    user_id = request.session["user_id"]

    resumes = Resume.objects.filter(
        user_id=user_id
    ).order_by("-uploaded_at")

    

    wb = Workbook()
    ws = wb.active
    ws.title = "Resume History"

    headers = [
        "Resume",
        "Upload Date",
        "ATS Score",
        "Skills Found",
        "Missing Skills",
        "AI Suggestion",
    ]

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="2563EB"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
        size=12
    )

    thin = Side(
        border_style="thin",
        color="CCCCCC"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for col, header in enumerate(headers, 1):

        cell = ws.cell(row=1, column=col)

        cell.value = header

        cell.fill = header_fill

        cell.font = header_font

        cell.border = border

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    row = 2

    for resume in resumes:

        ws.cell(row=row, column=1).value = resume.resume.name

        ws.cell(row=row, column=2).value = (
            resume.uploaded_at.strftime("%d-%m-%Y %I:%M %p")
            if resume.uploaded_at else ""
        )

        ws.cell(row=row, column=3).value = f"{resume.ats_score}%"

        ws.cell(row=row, column=4).value = resume.skills_found or "-"

        ws.cell(row=row, column=5).value = resume.missing_skills or "-"

        ws.cell(row=row, column=6).value = resume.ai_suggestion or "-"

        for col in range(1, 7):

            c = ws.cell(row=row, column=col)

            c.border = border

            c.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

        row += 1

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:

        length = 0

        letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:

            try:

                if cell.value:

                    length = max(
                        length,
                        len(str(cell.value))
                    )

            except:

                pass

        ws.column_dimensions[letter].width = min(length + 5, 60)

    ws.row_dimensions[1].height = 25

    wb.properties.creator = "AI Resume Screening System"

    wb.properties.title = "Resume History"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="resume_history.xlsx"'
    )

    wb.save(response)

    return response
# ===========================
# Download Report
# ===========================

def download_report(request, id):

    if "user_id" not in request.session:
        return redirect("login")

    user = Register.objects.get(id=request.session["user_id"])
    resume = get_object_or_404(
        Resume,
        id=id,
        user=user
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="ATS_Report.pdf"'

    p = canvas.Canvas(response)

    p.setFont("Helvetica-Bold", 18)
    p.drawString(180, 800, "AI Resume Screening Report")

    p.setFont("Helvetica", 12)

    y = 760

    p.drawString(50, y, f"Candidate : {user.fullname}")
    y -= 25

    p.drawString(50, y, f"Email : {user.email}")
    y -= 25

    p.drawString(50, y, f"ATS Score : {resume.ats_score}%")
    y -= 25

    p.drawString(50, y, "Skills Found:")
    y -= 20
    p.drawString(70, y, resume.skills_found)
    y -= 40

    p.drawString(50, y, "Missing Skills:")
    y -= 20
    p.drawString(70, y, resume.missing_skills)
    y -= 40

    p.drawString(50, y, "AI Suggestions:")
    y -= 20
    p.drawString(70, y, resume.ai_suggestion)
    y -= 60

    p.drawString(50, y, f"Uploaded On : {resume.uploaded_at}")

    y -= 60

    p.setFont("Helvetica-Bold", 14)
    p.drawString(120, y, "Generated By AI Resume Screening System")

    p.showPage()
    p.save()

    return response

def delete_resume(request, id):

    if "user_id" not in request.session:
        return redirect("login")

    user = Register.objects.get(id=request.session["user_id"])

    resume = get_object_or_404(
        Resume,
        id=id,
        user=user
    )

    if resume.resume:
        resume.resume.delete(save=False)

    resume.delete()
    messages.success(request, "Resume Deleted Successfully.")

    return redirect("history")

# ===========================
# Profile
# ===========================

def profile(request):

    if "user_id" not in request.session:
        return redirect("login")

    user = Register.objects.get(id=request.session["user_id"])

    resumes = Resume.objects.filter(user=user)

    total_resume = resumes.count()

    highest_score = 0
    latest_score = 0
    average_score = 0

    if resumes.exists():

        highest_score = resumes.order_by(
            "-ats_score"
        ).first().ats_score

        latest_score = resumes.order_by(
            "-uploaded_at"
        ).first().ats_score

        average_score = round(
            resumes.aggregate(
                avg=Avg("ats_score")
            )["avg"] or 0
        )

    context = {

        "user": user,

        "total_resume": total_resume,

        "highest_score": highest_score,

        "latest_score": latest_score,

        "average_score": average_score,

    }

    return render(
        request,
        "profile.html",
        context
    )
# ===========================
# Forgot Password
# ===========================

def forgot_password(request):

    if request.method == "POST":

        email = request.POST.get("email")

        user = Register.objects.filter(email=email).first()

        if not user:

            messages.error(request, "No account found with this email.")

            return render(request, "forgot_password.html", {
                "error": "No account found with this email."
            })

        # Store user id in session
        request.session["reset_user_id"] = user.id

        return redirect("change_password")

    return render(request, "forgot_password.html")

# ===========================
# Change Password
# ===========================

def change_password(request):

    if "reset_user_id" not in request.session:
        return redirect("forgot_password")

    user = Register.objects.get(
        id=request.session["reset_user_id"]
    )

    if request.method == "POST":

        password = request.POST.get(
            "password",
            ""
        )

        confirm = request.POST.get(
            "confirm_password",
            ""
        )

        if not password:
            messages.error(
                request,
                "Password is required."
            )
            return render(
                request,
                "change_password.html"
            )

        if len(password) < 8:
            messages.error(
                request,
                "Password must be at least 8 characters."
            )
            return render(
                request,
                "change_password.html"
            )

        if password != confirm:
            messages.error(
                request,
                "Passwords do not match."
            )
            return render(
                request,
                "change_password.html"
            )

        user.password = make_password(password)

        user.save()

        request.session.pop(
            "reset_user_id",
            None
        )

        messages.success(
            request,
            "Password Changed Successfully."
        )

        return redirect("login")

    return render(
        request,
        "change_password.html"
    )
# ===========================
# Analytics
# ===========================

def analytics(request):

    if "user_id" not in request.session:
        return redirect("login")

    user = Register.objects.get(id=request.session["user_id"])

    resumes = Resume.objects.filter(user=user)

    total_resume = resumes.count()

    highest_score = 0
    average_score = 0
    latest_score = 0
    lowest_score = 0
    ai_rating = "No Resume"

    if resumes.exists():

        scores = [r.ats_score for r in resumes]

        highest_score = max(scores)

        lowest_score = min(scores)

        average_score = round(sum(scores) / len(scores), 2)

        latest_score = resumes.order_by("-uploaded_at").first().ats_score

        if average_score >= 80:
            ai_rating = "Excellent ⭐⭐⭐⭐⭐"

        elif average_score >= 60:
            ai_rating = "Good ⭐⭐⭐⭐"

        elif average_score >= 40:
            ai_rating = "Average ⭐⭐⭐"

        else:
            ai_rating = "Needs Improvement ⭐⭐"

    context = {

        "user": user,

        "total_resume": total_resume,

        "highest_score": highest_score,

        "lowest_score": lowest_score,

        "latest_score": latest_score,

        "average_score": average_score,

        "ai_rating": ai_rating,

    }

    return render(request, "analytics.html", context)
# ===========================
# Contact
# ===========================

def contact(request):

    if "user_id" not in request.session:
        return redirect("login")

    user = Register.objects.get(
        id=request.session["user_id"]
    )

    if request.method == "POST":

        name = request.POST.get(
            "name",
            ""
        ).strip()

        email = request.POST.get(
            "email",
            ""
        ).strip()

        subject = request.POST.get(
            "subject",
            ""
        ).strip()

        message = request.POST.get(
            "message",
            ""
        ).strip()

        if not name:
            messages.error(
                request,
                "Name is required."
            )
            return redirect("contact")

        if not email:
            messages.error(
                request,
                "Email is required."
            )
            return redirect("contact")

        if not subject:
            messages.error(
                request,
                "Subject is required."
            )
            return redirect("contact")

        if not message:
            messages.error(
                request,
                "Message is required."
            )
            return redirect("contact")

        Contact.objects.create(

            user=user,

            name=name,

            email=email,

            subject=subject,

            message=message,

        )

        messages.success(
            request,
            "Message Sent Successfully."
        )

        return redirect("contact")

    return render(
        request,
        "contact.html",
        {
            "user": user
        }
    )
    
    # ===========================
# Edit Profile
# ===========================



def edit_profile(request):

    if "user_id" not in request.session:
        return redirect("login")

    user = Register.objects.get(id=request.session["user_id"])

    resumes = Resume.objects.filter(user=user)

    total_resume = resumes.count()

    highest_score = 0
    average_score = 0

    if resumes.exists():

        highest_score = resumes.order_by(
            "-ats_score"
        ).first().ats_score

        average_score = round(
            resumes.aggregate(
                avg=Avg("ats_score")
            )["avg"] or 0
        )

    if request.method == "POST":

        fullname = request.POST.get("fullname", "").strip()
        email = request.POST.get("email", "").strip().lower()

        profile_photo = request.FILES.get("profile_photo")

        if profile_photo:

            allowed = ["jpg", "jpeg", "png"]

            ext = profile_photo.name.split(".")[-1].lower()

            if ext not in allowed:

                messages.error(
                    request,
                    "Only JPG, JPEG and PNG images are allowed."
                )

                return redirect("edit_profile")

        if Register.objects.filter(
            email=email
        ).exclude(id=user.id).exists():

            messages.error(
                request,
                "Email already exists."
            )

            return render(
                request,
                "edit_profile.html",
                {
                    "user": user,
                    "total_resume": total_resume,
                    "highest_score": highest_score,
                    "average_score": average_score,
                }
            )

        user.fullname = fullname
        user.email = email

        if profile_photo:
            user.profile_photo = profile_photo

        user.save()

        request.session["user_name"] = user.fullname

        messages.success(
            request,
            "Profile Updated Successfully."
        )

        return redirect("edit_profile")

    return render(
        request,
        "edit_profile.html",
        {
            "user": user,
            "total_resume": total_resume,
            "highest_score": highest_score,
            "average_score": average_score,
        }
    )

# ==========================================================
# ADMIN DASHBOARD
# ==========================================================

def admin_dashboard(request):

    if not admin_required(request):
        return redirect("admin_login")

    admin = get_object_or_404(
        Register,
        id=request.session["user_id"]
    )

    # ======================================================
    # Dashboard Statistics
    # ======================================================

    total_users = Register.objects.filter(
        is_admin=False
    ).count()

    total_resumes = Resume.objects.count()

    total_messages = Contact.objects.count()

    ai_reviews = total_resumes

    pending_reviews = Resume.objects.filter(
        ats_score=0
    ).count()

    average_score = round(

        Resume.objects.aggregate(
            Avg("ats_score")
        )["ats_score__avg"] or 0,

        2

    )

    highest_score = (

        Resume.objects.aggregate(
            Max("ats_score")
        )["ats_score__max"] or 0

    )

    lowest_score = (

        Resume.objects.aggregate(
            Min("ats_score")
        )["ats_score__min"] or 0

    )

    active_users = (

        Resume.objects.values(
            "user"
        ).distinct().count()

    )

    # ======================================================
    # Recent Tables
    # ======================================================

    recent_resumes = (

        Resume.objects

        .select_related("user")

        .order_by("-uploaded_at")[:10]

    )

    recent_users = (

        Register.objects

        .filter(is_admin=False)

        .order_by("-created_at")[:10]

    )

    top_scores = (

        Resume.objects

        .select_related("user")

        .order_by("-ats_score")[:10]

    )

    # ======================================================
    # Progress Cards
    # ======================================================

    processed_percent = 100 if total_resumes else 0

    suggestion_percent = 92

    accuracy_percent = 97

    # ======================================================
    # ATS Distribution
    # ======================================================

    excellent_count = Resume.objects.filter(
        ats_score__gte=80
    ).count()

    good_count = Resume.objects.filter(
        ats_score__gte=60,
        ats_score__lt=80
    ).count()

    average_count = Resume.objects.filter(
        ats_score__gte=40,
        ats_score__lt=60
    ).count()

    poor_count = Resume.objects.filter(
        ats_score__lt=40
    ).count()

    # ======================================================
    # Monthly Upload Report
    # ======================================================

    monthly_uploads = (

        Resume.objects

        .annotate(
            month=ExtractMonth("uploaded_at")
        )

        .values("month")

        .annotate(
            total=Count("id")
        )

        .order_by("month")

    )

    month_data = {
        i: 0 for i in range(1, 13)
    }

    for item in monthly_uploads:

        month_data[item["month"]] = item["total"]
        month_labels = [

        "Jan", "Feb", "Mar", "Apr",
        "May", "Jun", "Jul", "Aug",
        "Sep", "Oct", "Nov", "Dec"

    ]
        for item in monthly_uploads:

            month_data[item["month"]] = item["total"]

    month_values = [

        month_data[1],
        month_data[2],
        month_data[3],
        month_data[4],
        month_data[5],
        month_data[6],
        month_data[7],
        month_data[8],
        month_data[9],
        month_data[10],
        month_data[11],
        month_data[12],

    ]

    # ======================================================
    # Top Skills
    # ======================================================

    skill_counter = {}

    resumes = Resume.objects.exclude(
        skills_found__isnull=True
    ).exclude(
        skills_found=""
    )

    for resume in resumes:

        skills = [
            skill.strip()
            for skill in resume.skills_found.split(",")
            if skill.strip()
        ]

        for skill in skills:

            skill_counter[skill] = (

                skill_counter.get(skill, 0) + 1

            )

    top_skills = []

    for skill, count in sorted(

        skill_counter.items(),

        key=lambda x: x[1],

        reverse=True

    )[:10]:

        top_skills.append({

            "name": skill,

            "count": count

        })

    # ======================================================
    # Recent Activities
    # ======================================================

    recent_activities = []

    for resume in recent_resumes:

        recent_activities.append({

            "title": f"{resume.user.fullname} uploaded a resume",

            "description": f"ATS Score : {resume.ats_score}%",

            "created_at": resume.uploaded_at,

        })

    # ======================================================
    # Chart Data
    # ======================================================

    chart_labels = month_labels

    chart_values = month_values

    pie_chart = {

        "excellent": excellent_count,

        "good": good_count,

        "average": average_count,

        "poor": poor_count,

    }
    
    
# ======================================================
# Context
# ======================================================

    context = {

        "admin": admin,

        # Statistics
        "total_users": total_users,
        "total_resumes": total_resumes,
        "total_messages": total_messages,
        "active_users": active_users,

        # ATS
        "average_score": average_score,
        "highest_score": highest_score,
        "lowest_score": lowest_score,

        # AI
        "ai_reviews": ai_reviews,
        "pending_reviews": pending_reviews,

        # Progress
        "processed_percent": processed_percent,
        "suggestion_percent": suggestion_percent,
        "accuracy_percent": accuracy_percent,

        # Tables
        "recent_resumes": recent_resumes,
        "recent_users": recent_users,
        "top_scores": top_scores,

        # Skills
        "top_skills": top_skills,

        # Activity
        "recent_activities": recent_activities,

        # Charts
        "chart_labels": chart_labels,
        "chart_values": chart_values,
        "month_labels": month_labels,
        "month_values": month_values,

        # Pie Chart
        "excellent_count": excellent_count,
        "good_count": good_count,
        "average_count": average_count,
        "poor_count": poor_count,

        "pie_chart": pie_chart,

    }

    return render(

        request,

        "admin/admin_dashboard.html",

        context

    )

# ==========================================================
# MANAGE USERS
# ==========================================================


def manage_users(request):

    if not admin_required(request):
        return redirect("admin_login")

    search = request.GET.get("search", "").strip()
    role = request.GET.get("role", "").strip()

    users = Register.objects.all().order_by("-id")

    if search:
        users = users.filter(
            Q(fullname__icontains=search) |
            Q(email__icontains=search)
        )

    if role == "admin":
        users = users.filter(is_admin=True)

    elif role == "user":
        users = users.filter(is_admin=False)

    paginator = Paginator(users, 10)

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(page_number)

    context = {

        "users": page_obj,

        "page_obj": page_obj,

        "search": search,

        "role": role,

        "total_users": Register.objects.count(),

        "admin_count": Register.objects.filter(
            is_admin=True
        ).count(),

        "active_users": Resume.objects.values(
            "user"
        ).distinct().count(),

        "today_users": Register.objects.filter(
            created_at__date=timezone.now().date()
        ).count(),

    }

    return render(
        request,
        "admin/manage_users.html",
        context
    )


# ==========================================================
# MANAGE RESUME
# ==========================================================

def manage_resume(request):

    if not admin_required(request):
        return redirect("admin_login")

    search = request.GET.get("search", "").strip()
    score = request.GET.get("score", "").strip()

    resumes = Resume.objects.select_related(
        "user"
    ).order_by("-uploaded_at")

    if search:

        resumes = resumes.filter(

            Q(user__fullname__icontains=search) |

            Q(user__email__icontains=search) |

            Q(skills_found__icontains=search) |

            Q(missing_skills__icontains=search)

        )

    if score:

        try:

            resumes = resumes.filter(
                ats_score__gte=int(score)
            )

        except ValueError:
            pass

    paginator = Paginator(resumes, 10)

    page_number = request.GET.get("page")

    resumes = paginator.get_page(page_number)

    total_resume = Resume.objects.count()

    average_score = round(

        Resume.objects.aggregate(
            Avg("ats_score")
        )["ats_score__avg"] or 0

    )

    highest_score = Resume.objects.aggregate(
        Max("ats_score")
    )["ats_score__max"] or 0

    excellent_count = Resume.objects.filter(
        ats_score__gte=90
    ).count()

    context = {

        "resumes": resumes,

        "search": search,

        "score": score,

        "total_resumes": total_resume,

        "average_score": average_score,

        "highest_score": highest_score,

        "excellent_count": excellent_count,

    }

    return render(

        request,

        "admin/manage_resume.html",

        context

    )

# ==========================================================
# CONTACT MESSAGES
# ==========================================================

def contact_messages(request):

    if not admin_required(request):
        return redirect("admin_login")

    search = request.GET.get("search", "").strip()
    status = request.GET.get("status", "").strip().lower()

    messages_qs = (
        Contact.objects
        .select_related("user")
        .order_by("-created_at")
    )

    if search:
        messages_qs = messages_qs.filter(
            Q(name__icontains=search) |
            Q(email__icontains=search) |
            Q(subject__icontains=search) |
            Q(message__icontains=search)
        )

    # Read / Unread Support
    has_is_read = hasattr(Contact, "is_read")

    if has_is_read:

        if status == "read":
            messages_qs = messages_qs.filter(is_read=True)

        elif status == "unread":
            messages_qs = messages_qs.filter(is_read=False)

        read_messages = Contact.objects.filter(
            is_read=True
        ).count()

        unread_messages = Contact.objects.filter(
            is_read=False
        ).count()

    else:

        read_messages = 0

        unread_messages = Contact.objects.count()

    deleted_messages = 0

    paginator = Paginator(messages_qs, 10)

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(page_number)

    context = {

        "messages": page_obj,

        "page_obj": page_obj,

        "search": search,

        "status": status,

        "total_messages": Contact.objects.count(),

        "read_messages": read_messages,

        "unread_messages": unread_messages,

        "deleted_messages": deleted_messages,

    }

    return render(

        request,

        "admin/contact_messages.html",

        context

    )


# ==========================================================
# DELETE MESSAGE
# ==========================================================

def delete_message(request, id):

    if not admin_required(request):
        return redirect("admin_login")

    contact = get_object_or_404(
        Contact,
        id=id
    )

    contact.delete()

    messages.success(
        request,
        "Message Deleted Successfully."
    )

    return redirect("contact_messages")

# ==========================================================
# ADMIN ANALYTICS
# ==========================================================

def admin_analytics(request):

    if not admin_required(request):
        return redirect("admin_login")

    total_users = Register.objects.filter(
        is_admin=False
    ).count()

    total_resumes = Resume.objects.count()

    total_messages = Contact.objects.count()

    average_score = round(
        Resume.objects.aggregate(
            Avg("ats_score")
        )["ats_score__avg"] or 0,
        2
    )

    highest_score = (
        Resume.objects.aggregate(
            Max("ats_score")
        )["ats_score__max"] or 0
    )

    lowest_score = (
        Resume.objects.aggregate(
            Min("ats_score")
        )["ats_score__min"] or 0
    )

    active_users = (
        Resume.objects.values("user")
        .distinct()
        .count()
    )

    monthly_uploads = (
        Resume.objects
        .annotate(
            month=ExtractMonth("uploaded_at")
        )
        .values("month")
        .annotate(total=Count("id"))
        .order_by("month")
    )

    month_data = {i: 0 for i in range(1, 13)}

    for item in monthly_uploads:
        month_data[item["month"]] = item["total"]

    monthly_labels = [
        "Jan", "Feb", "Mar", "Apr",
        "May", "Jun", "Jul", "Aug",
        "Sep", "Oct", "Nov", "Dec"
    ]

    monthly_values = [
        month_data[1],
        month_data[2],
        month_data[3],
        month_data[4],
        month_data[5],
        month_data[6],
        month_data[7],
        month_data[8],
        month_data[9],
        month_data[10],
        month_data[11],
        month_data[12],
    ]

    score_distribution = {

        "excellent": Resume.objects.filter(
            ats_score__gte=80
        ).count(),

        "good": Resume.objects.filter(
            ats_score__gte=60,
            ats_score__lt=80
        ).count(),

        "average": Resume.objects.filter(
            ats_score__gte=40,
            ats_score__lt=60
        ).count(),

        "poor": Resume.objects.filter(
            ats_score__lt=40
        ).count(),

    }

    recent_uploads = Resume.objects.select_related(
        "user"
    ).order_by("-uploaded_at")[:10]

    top_users = (
        Register.objects.filter(is_admin=False)
        .annotate(
            resume_count=Count("resumes")
        )
        .order_by("-resume_count")[:10]
    )
    # -----------------------------
    # Top Skills
    # -----------------------------

    skill_counter = {}

    for resume in Resume.objects.exclude(
        skills_found__isnull=True
    ).exclude(
        skills_found=""
    ):

        for skill in resume.skills_found.split(","):

            skill = skill.strip()

            if skill:

                skill_counter[skill] = (
                    skill_counter.get(skill, 0) + 1
                )

    top_skills = sorted(

        skill_counter.items(),

        key=lambda x: x[1],

        reverse=True

    )[:10]

    skill_labels = [item[0] for item in top_skills]

    skill_values = [item[1] for item in top_skills]

    chart_data = {

        "monthly_labels": monthly_labels,

        "monthly_values": monthly_values,

        "score_distribution": score_distribution,

        "skill_labels": skill_labels,

        "skill_values": skill_values,

    }

    export_report = {

        "generated_by": "AI Resume Screening System",

        "total_users": total_users,

        "total_resumes": total_resumes,

        "average_score": average_score,

        "highest_score": highest_score,

        "lowest_score": lowest_score,

        "generated_at": Resume.objects.order_by(
            "-uploaded_at"
        ).first(),

    }

    context = {

        "total_users": total_users,

        "total_resumes": total_resumes,

        "total_messages": total_messages,

        "average_score": average_score,

        "highest_score": highest_score,

        "lowest_score": lowest_score,

        "active_users": active_users,

        "recent_uploads": recent_uploads,

        "top_users": top_users,

        "top_skills": top_skills,

        "monthly_labels": monthly_labels,

        "monthly_values": monthly_values,

        "score_distribution": score_distribution,

        "chart_data": chart_data,

        "export_report": export_report,

        "skill_labels": skill_labels,

        "skill_values": skill_values,

    }

    return render(
        request,
        "admin/admin_analytics.html",
        context
    )


# =====================================================
# ADMIN LOGIN
# =====================================================

def admin_login(request):

    # Already Logged In
    if request.session.get("is_admin"):
        return redirect("admin_dashboard")

    if request.method == "POST":

        email = request.POST.get("email", "").strip().lower()
        password = request.POST.get("password", "")

        if not email or not password:
            messages.error(request, "Email and Password are required.")
            return render(request, "admin/admin_login.html")

        admin = Register.objects.filter(
            email=email,
            is_admin=True
        ).first()
        
        print("ADMIN =", admin)

        if not admin:
            print("ADMIN NOT FOUND")
            print("EMAIL =", email)
            
            
            messages.error(request, "Invalid Admin Email or Password.")
            return render(request, "admin/admin_login.html")
        
        print("ADMIN PASSWORD =", admin.password)
        print("CHECK =", check_password(password, admin.password))

        if not check_password(password, admin.password):
            messages.error(request, "Invalid Admin Email or Password.")
            return render(request, "admin/admin_login.html")

        request.session["user_id"] = admin.id
        request.session["user_name"] = admin.fullname
        request.session["is_admin"] = True

        messages.success(request, "Admin Login Successful.")

        return redirect("admin_dashboard")

    return render(request, "admin/admin_login.html")
    
# =====================================================
# ADMIN LOGOUT
# =====================================================

def admin_logout(request):

    request.session.flush()

    return redirect("admin_login")

# =====================================================
# ADMIN CHECK
# =====================================================

def admin_required(request):

    if "user_id" not in request.session:

        return False

    if not request.session.get("is_admin"):

        return False

    return True

def view_user(request, id):

    if not admin_required(request):
        return redirect("admin_login")

    user = get_object_or_404(Register, id=id)

    return render(
        request,
        "admin/view_user.html",
        {
            "user": user
        }
    )


def edit_user(request, id):

    if not admin_required(request):
        return redirect("admin_login")

    user = get_object_or_404(Register, id=id)

    if request.method == "POST":

        user.fullname = request.POST.get("fullname")

        user.email = request.POST.get("email")

        user.is_admin = request.POST.get("is_admin") == "on"

        user.save()

        messages.success(request, "User Updated Successfully.")

        return redirect("manage_users")

    return render(
        request,
        "admin/edit_user.html",
        {
            "user": user
        }
    )

def delete_user(request, id):

    if not admin_required(request):
        return redirect("admin_login")

    user = get_object_or_404(Register, id=id)

    # Cannot delete your own account
    if user.id == request.session["user_id"]:

        messages.error(
            request,
            "You cannot delete your own account."
        )

        return redirect("manage_users")

    # Cannot delete another admin
    if user.is_admin:

        messages.error(
            request,
            "Admin account cannot be deleted."
        )

        return redirect("manage_users")

    user.delete()

    messages.success(
        request,
        "User Deleted Successfully."
    )

    return redirect("manage_users")
# ==========================================================
# DELETE RESUME (ADMIN)
# ==========================================================

def delete_resume_admin(request, id):

    if not admin_required(request):
        return redirect("admin_login")

    resume = get_object_or_404(Resume, id=id)

    if resume.resume:
        resume.resume.delete(save=False)

    resume.delete()

    messages.success(request, "Resume Deleted Successfully.")

    return redirect("manage_resume")

# ==========================================================
# TOGGLE ADMIN
# ==========================================================

def toggle_admin(request, id):

    if not admin_required(request):
        return redirect("admin_login")

    user = get_object_or_404(
        Register,
        id=id
    )

    current_admin = Register.objects.get(
        id=request.session["user_id"]
    )

    if user.id == current_admin.id:

        messages.error(
            request,
            "You cannot change your own admin status."
        )

        return redirect("manage_users")

    user.is_admin = not user.is_admin

    user.save()

    if user.is_admin:

        messages.success(
            request,
            f"{user.fullname} is now an Admin."
        )

    else:

        messages.success(
            request,
            f"{user.fullname} is now a Normal User."
        )

    return redirect("manage_users")


def admin_export_excel(request):

    if "user_id" not in request.session or not request.session.get("is_admin"):
        return redirect("admin_login")

    resumes = Resume.objects.select_related("user").all().order_by("-uploaded_at")

    search = request.GET.get("search", "").strip()
    score = request.GET.get("score", "").strip()

    if search:
        resumes = resumes.filter(
            Q(user__fullname__icontains=search) |
            Q(user__email__icontains=search) |
            Q(skills_found__icontains=search) |
            Q(missing_skills__icontains=search) |
            Q(ai_suggestion__icontains=search)
        )

    if score:
        try:
            resumes = resumes.filter(ats_score=int(score))
        except ValueError:
            pass

    wb = Workbook()
    ws = wb.active
    ws.title = "Manage Resume"

    headers = [
        "ID",
        "Candidate Name",
        "Email",
        "Resume",
        "Upload Date",
        "ATS Score",
        "Skills Found",
        "Missing Skills",
        "AI Suggestion",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 2

    for resume in resumes:

        ws.cell(row=row, column=1).value = resume.id
        ws.cell(row=row, column=2).value = resume.user.fullname
        ws.cell(row=row, column=3).value = resume.user.email
        ws.cell(row=row, column=4).value = resume.resume.name
        ws.cell(row=row, column=5).value = resume.uploaded_at.strftime("%d-%m-%Y %I:%M %p")
        ws.cell(row=row, column=6).value = f"{resume.ats_score}%"
        ws.cell(row=row, column=7).value = resume.skills_found or ""
        ws.cell(row=row, column=8).value = resume.missing_skills or ""
        ws.cell(row=row, column=9).value = resume.ai_suggestion or ""

        row += 1

    # Auto Width
    for column_cells in ws.columns:
        max_length = 0

        column = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        ws.column_dimensions[column].width = max_length + 5

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="admin_resume_report.xlsx"'
    )

    wb.save(response)

    return response

def admin_export_pdf(request):

    if "user_id" not in request.session or not request.session.get("is_admin"):
        return redirect("admin_login")

    resumes = Resume.objects.select_related("user").all().order_by("-uploaded_at")

    search = request.GET.get("search", "").strip()
    score = request.GET.get("score", "").strip()

    if search:
        resumes = resumes.filter(
            Q(user__fullname__icontains=search) |
            Q(user__email__icontains=search) |
            Q(skills_found__icontains=search) |
            Q(missing_skills__icontains=search) |
            Q(ai_suggestion__icontains=search)
        )

    if score:
        try:
            resumes = resumes.filter(ats_score=int(score))
        except ValueError:
            pass

    response = HttpResponse(
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        'attachment; filename="admin_resume_report.pdf"'
    )

    pdf = canvas.Canvas(response)

    text = pdf.beginText()

    text.setTextOrigin(40, 800)

    text.setFont("Helvetica-Bold", 16)
    text.textLine("AI Resume Screening System")
    text.textLine("Admin Resume Report")
    text.textLine("")

    text.setFont("Helvetica", 10)

    for resume in resumes:

        text.textLine(f"Resume ID : {resume.id}")
        text.textLine(f"Candidate : {resume.user.fullname}")
        text.textLine(f"Email     : {resume.user.email}")
        text.textLine(f"Resume    : {resume.resume.name}")
        text.textLine(
            f"Uploaded  : {resume.uploaded_at.strftime('%d-%m-%Y %I:%M %p')}"
        )
        text.textLine(f"ATS Score : {resume.ats_score}%")
        text.textLine(f"Skills    : {resume.skills_found or '-'}")
        text.textLine(f"Missing   : {resume.missing_skills or '-'}")
        text.textLine(f"Suggestion: {resume.ai_suggestion or '-'}")
        text.textLine("-" * 100)

        if text.getY() < 80:
            pdf.drawText(text)
            pdf.showPage()
            text = pdf.beginText()
            text.setTextOrigin(40, 800)
            text.setFont("Helvetica", 10)

    pdf.drawText(text)
    pdf.save()

    return response




def debug_admin(request):
    data = list(
        Register.objects.values(
            "id",
            "fullname",
            "email",
            "is_admin"
        )
    )
    return JsonResponse(data, safe=False)

def error_404(request, exception):
    return render(request, "404.html", status=404)


def error_500(request):
    return render(request, "500.html", status=500)